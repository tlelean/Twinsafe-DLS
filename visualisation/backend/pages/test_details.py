from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel
from fastapi.responses import JSONResponse, FileResponse
from pathlib import Path

import base64
import binascii
import os
import tempfile
import threading
import platform
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ..opc import opc

router = APIRouter()

DETAILS_DIR = Path("/mnt/otsdls/Details")
TEMPLATES_DIR = Path("/mnt/otsdls/Test Log Sheet Templates")

# ----------------------------
# Existing .bin generation bits
# ----------------------------

BIN_FIELD_ORDER = ["OTS Number", "Drawing Number", "Client", "Line Item", "User"]
BIN_FIELD_SIZE = 81

_bin_gen_lock = threading.Lock()


def _to_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _pack_field(text: str) -> bytes:
    raw = _to_text(text).encode("utf-8", errors="ignore")
    return raw[:BIN_FIELD_SIZE].ljust(BIN_FIELD_SIZE, b"\x00")


def _build_bin(values: Dict[str, str]) -> bytes:
    names_block = b"".join(_pack_field(k) for k in BIN_FIELD_ORDER)
    values_block = b"".join(_pack_field(values.get(k, "")) for k in BIN_FIELD_ORDER)
    return names_block + values_block


def _set_hidden_windows(path: Path) -> None:
    if platform.system() != "Windows":
        return
    try:
        os.system(f'attrib +h "{path}"')
    except Exception:
        pass


def _atomic_write(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.NamedTemporaryFile(
        mode="wb",
        delete=False,
        dir=str(path.parent),
        prefix=".tmp_",
        suffix=path.suffix,
    ) as tmp:
        tmp.write(data)
        tmp.flush()
        os.fsync(tmp.fileno())
        tmp_name = tmp.name

    os.replace(tmp_name, str(path))

    _set_hidden_windows(path)
    _set_hidden_windows(path.parent)


def _generate_bins_from_templates() -> None:
    if not TEMPLATES_DIR.exists():
        return

    DETAILS_DIR.mkdir(parents=True, exist_ok=True)
    _set_hidden_windows(DETAILS_DIR)

    for xlsx_path in TEMPLATES_DIR.glob("*.xlsx"):
        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)

            if "Test Info" not in wb.sheetnames:
                continue

            ws = wb["Test Info"]

            ots_number = _to_text(ws["C9"].value)
            drawing_no = _to_text(ws["C11"].value)
            client = _to_text(ws["C3"].value)
            line_item = _to_text(ws["C10"].value)

            values = {
                "OTS Number": ots_number,
                "Drawing Number": drawing_no,
                "Client": client,
                "Line Item": line_item,
                "User": "",
            }

            out_name = f"{ots_number}.bin" if ots_number else f"{xlsx_path.stem}.bin"
            out_path = (DETAILS_DIR / out_name).resolve()

            if DETAILS_DIR.resolve() not in out_path.parents:
                continue

            if out_path.exists() and out_path.stat().st_mtime >= xlsx_path.stat().st_mtime:
                continue

            _atomic_write(out_path, _build_bin(values))

        except Exception:
            continue


@router.get("/api/details")
def list_details_bins():
    try:
        with _bin_gen_lock:
            _generate_bins_from_templates()

        DETAILS_DIR.mkdir(parents=True, exist_ok=True)
        _set_hidden_windows(DETAILS_DIR)

        files = [p for p in DETAILS_DIR.iterdir() if p.is_file() and p.suffix.lower() == ".bin"]
        files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return [p.name for p in files]

    except Exception:
        return JSONResponse(status_code=500, content={"error": "Failed to read /mnt/otsdls/Details"})


class DetailsSelection(BaseModel):
    filename: str


@router.post("/api/details/select")
def select_details_file(payload: DetailsSelection):
    filename = payload.filename

    if "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    if not filename.lower().endswith(".bin"):
        raise HTTPException(status_code=400, detail="Only .bin files allowed")

    file_path = (DETAILS_DIR / filename).resolve()
    if DETAILS_DIR.resolve() not in file_path.parents:
        raise HTTPException(status_code=400, detail="Invalid path")

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    try:
        opc.write("filename_details_production", filename)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OPC write failed: {e}")

    return {"ok": True, "filename": filename}

@router.get("/api/details/ots")
def get_current_ots_number():
    try:
        ots = opc.read("ots_number")
    except KeyError:
        raise HTTPException(status_code=500, detail="OPC node missing 'ots_number_production'")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OPC read failed: {e}")

    return {"ots_number": str(ots).strip() if ots is not None else ""}

@router.get("/api/details/file/{filename}")
def get_details_file(filename: str):
    if "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    if not filename.lower().endswith(".bin"):
        raise HTTPException(status_code=400, detail="Only .bin files allowed")

    file_path = (DETAILS_DIR / filename).resolve()
    if DETAILS_DIR.resolve() not in file_path.parents:
        raise HTTPException(status_code=400, detail="Invalid path")

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(path=str(file_path), media_type="application/octet-stream", filename=file_path.name)


class DetailsSave(BaseModel):
    filename: str
    bytes_base64: str


@router.post("/api/details/save")
def save_details_file(payload: DetailsSave):
    filename = payload.filename

    if "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    if not filename.lower().endswith(".bin"):
        raise HTTPException(status_code=400, detail="Only .bin files allowed")

    try:
        raw = base64.b64decode(payload.bytes_base64, validate=True)
    except (binascii.Error, ValueError):
        raise HTTPException(status_code=400, detail="Invalid base64 payload")

    out_path = (DETAILS_DIR / filename).resolve()
    if DETAILS_DIR.resolve() not in out_path.parents:
        raise HTTPException(status_code=400, detail="Invalid path")

    try:
        _atomic_write(out_path, raw)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {e}")

    return {"ok": True, "filename": out_path.name}


# -----------------------------------------
# NEW: Log sheet extraction + hosted endpoints
# -----------------------------------------

LOG_ROWS = range(10, 28)  # 10 to 27 inclusive
LOG_COLS = {
    "section_number": "A",
    "test_name": "B",
    "hold_time_minutes": "E",
    "test_pressure": "F",
}

_log_cache_lock = threading.Lock()
_log_cache_data: Dict[str, Dict] = {}          # ots -> {"source_files": [...], "sections": [...]}
_log_cache_mtimes: Dict[str, float] = {}       # filepath str -> mtime


def _safe_float(v) -> Optional[float]:
    """Try to turn Excel values into a float. Returns None if empty/unparseable."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _to_text(v)
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def _scan_templates_mtimes() -> Dict[str, float]:
    """Return mtimes for all .xlsx in templates dir."""
    mtimes: Dict[str, float] = {}
    if not TEMPLATES_DIR.exists():
        return mtimes
    for p in TEMPLATES_DIR.glob("*.xlsx"):
        try:
            mtimes[str(p.resolve())] = p.stat().st_mtime
        except Exception:
            continue
    return mtimes


def _cache_is_fresh(new_mtimes: Dict[str, float]) -> bool:
    """Cache is fresh if file set + mtimes match."""
    return new_mtimes == _log_cache_mtimes


def _load_logs_into_cache() -> None:
    """
    Builds cache grouped by OTS Number (Test Info!C9).
    From sheets 2..end (i.e. all sheets except 'Test Info'), extract:
      Rows 10..27, Cols A, B, E, F.
    """
    new_mtimes = _scan_templates_mtimes()
    if _cache_is_fresh(new_mtimes):
        return

    grouped: Dict[str, Dict] = {}

    for xlsx_path_str, _mtime in new_mtimes.items():
        xlsx_path = Path(xlsx_path_str)

        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)

            if "Test Info" not in wb.sheetnames:
                continue

            ws_info = wb["Test Info"]
            ots_number = _to_text(ws_info["C9"].value)

            if ots_number == "":
                # If there's no OTS Number, it can't be grouped for your UI
                continue

            # "Sheets 2 - end" => everything except "Test Info"
            data_sheet_names = [n for n in wb.sheetnames if n != "Test Info"]

            sections: List[Dict] = []

            for sheet_name in data_sheet_names:
                ws = wb[sheet_name]

                for r in LOG_ROWS:
                    section_number = _to_text(ws[f"{LOG_COLS['section_number']}{r}"].value)
                    test_name = _to_text(ws[f"{LOG_COLS['test_name']}{r}"].value)
                    hold_time = _safe_float(ws[f"{LOG_COLS['hold_time_minutes']}{r}"].value)
                    test_pressure = _safe_float(ws[f"{LOG_COLS['test_pressure']}{r}"].value)

                    # Skip rows where the section number is N/A (or NA)
                    sn = section_number.strip().lower()
                    if sn in ("n/a", "na") or "n/a" in sn:
                        continue

                    # Skip blank rows (no section number + no test name)
                    if section_number == "" and test_name == "":
                        continue

                    sections.append(
                        {
                            "section_number": section_number,
                            "test_name": test_name,
                            "hold_time_minutes": hold_time,
                            "test_pressure": test_pressure,
                            "sheet": sheet_name,
                            "template": xlsx_path.name,
                            "row": r,
                        }
                    )

            if ots_number not in grouped:
                grouped[ots_number] = {"source_files": [xlsx_path.name], "sections": []}
            else:
                grouped[ots_number]["source_files"].append(xlsx_path.name)

            grouped[ots_number]["sections"].extend(sections)

        except Exception:
            continue

    # Optional: sort each OTS group by section number then sheet then row (best-effort)
    def _sort_key(item: Dict) -> Tuple:
        # Try numeric sort on section number if possible
        s = item.get("section_number", "")
        try:
            n = float(s)
            return (0, n, item.get("sheet", ""), item.get("row", 0))
        except Exception:
            return (1, s, item.get("sheet", ""), item.get("row", 0))

    for ots in grouped:
        grouped[ots]["sections"].sort(key=_sort_key)

    _log_cache_data.clear()
    _log_cache_data.update(grouped)

    _log_cache_mtimes.clear()
    _log_cache_mtimes.update(new_mtimes)


@router.get("/api/logs/ots")
def list_ots_numbers():
    """
    Returns all OTS Numbers found in the templates folder.
    Use this to populate your OTS selection combobox.
    """
    try:
        with _log_cache_lock:
            _load_logs_into_cache()
        ots_list = sorted(_log_cache_data.keys())
        return {"ots_numbers": ots_list}
    except Exception:
        return JSONResponse(status_code=500, content={"error": "Failed to load log sheet templates"})


@router.get("/api/logs/{ots_number}/sections")
def list_sections_for_ots(
    ots_number: str,
    q: Optional[str] = Query(default=None, description="Type-ahead filter for section number (prefix match)"),
    limit: int = Query(default=50, ge=1, le=500),
):
    """
    Returns section list for a given OTS Number.
    For your combobox type-ahead:
      - call with q="12" and it will return sections whose section_number starts with "12".
    """
    try:
        with _log_cache_lock:
            _load_logs_into_cache()

        group = _log_cache_data.get(ots_number)
        if not group:
            raise HTTPException(status_code=404, detail="OTS Number not found")

        sections = group["sections"]

        if q is not None and q.strip() != "":
            qv = q.strip().lower()
            sections = [s for s in sections if _to_text(s.get("section_number", "")).lower().startswith(qv)]

        # Keep response light for type-ahead; return unique section numbers + a preview test name
        seen = set()
        out: List[Dict] = []
        for s in sections:
            sec = _to_text(s.get("section_number", ""))
            if sec == "" or sec in seen:
                continue
            seen.add(sec)
            out.append({"section_number": sec, "test_name": _to_text(s.get("test_name", ""))})
            if len(out) >= limit:
                break

        return {
            "ots_number": ots_number,
            "source_files": group.get("source_files", []),
            "sections": out,
        }

    except HTTPException:
        raise
    except Exception:
        return JSONResponse(status_code=500, content={"error": "Failed to read log sheet data"})


@router.get("/api/logs/{ots_number}/section/{section_number}")
def get_section_details(ots_number: str, section_number: str):
    """
    Returns the full details for one section:
      - test_name
      - hold_time_minutes
      - test_pressure
    If duplicates exist across sheets/templates, returns all matches as a list.
    """
    try:
        with _log_cache_lock:
            _load_logs_into_cache()

        group = _log_cache_data.get(ots_number)
        if not group:
            raise HTTPException(status_code=404, detail="OTS Number not found")

        target = section_number.strip().lower()

        matches = [
            s
            for s in group["sections"]
            if _to_text(s.get("section_number", "")).strip().lower() == target
        ]

        if not matches:
            raise HTTPException(status_code=404, detail="Section number not found")

        # If you expect unique section numbers, the UI can just take matches[0]
        return {"ots_number": ots_number, "section_number": section_number, "matches": matches}

    except HTTPException:
        raise
    except Exception:
        return JSONResponse(status_code=500, content={"error": "Failed to read section details"})